#####################################################
# フォルダ内のa&bと当日申し込みの2ファイルを１つのエクセルにまとめる
#####################################################

import pandas
import glob
import openpyxl

#ファイルの読み込み
read_path = 'C:\Tool_Wanpaku\受付名簿_2022年第46回わんぱく相撲渋谷区大会_ver1.0.xlsx'
output_path = "C:\Tool_Wanpaku\確定_全参加者リスト.xlsx"
wb = openpyxl.load_workbook(read_path)

#全シートのデータを取得する
sheet_names = wb.get_sheet_names()
first_row = 1
l_2d_output = []
for sheet_name in sheet_names:
    # もしsheet_nameが'a&b'か'当日申し込み'なら、sheet取得
    if sheet_name == 'a&b' or sheet_name == '当日申し込み':
        sheet = wb.get_sheet_by_name(sheet_name)

        for row in sheet.iter_rows(min_row=first_row):
            l = []
            for cell in row:
                l.append(str(cell.value))
            l_2d_output.append(l)

        #次シートからヘッダーの読み込みを飛ばす
        first_row = 2


#ファイルに書き込み
def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

#書き込み
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
write_list_2d(output_ws, l_2d_output, 1, 1)
output_wb.save(output_path)