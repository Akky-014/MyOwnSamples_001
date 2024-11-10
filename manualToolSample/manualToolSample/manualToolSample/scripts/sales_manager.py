import pandas as pd
from openpyxl import Workbook, load_workbook

# 1. Excelデータの読み込み（Pandasを使用）
input_file = "sales_data.xlsx"  # 入力ファイル名
df = pd.read_excel(input_file)

# 2. 月別の売上合計を計算
df['Month'] = pd.to_datetime(df['Date']).dt.to_period('M')  # 日付列を月単位に変換
monthly_sales = df.groupby('Month')['Sales'].sum().reset_index()

# 月別データを文字列型に変換
monthly_sales['Month'] = monthly_sales['Month'].astype(str)

# 3. OpenPyXLで元のExcelファイルに新しいシートを追加
wb = load_workbook(input_file)
ws = wb.create_sheet(title="Monthly Sales Summary")

# 4. 新しいシートにデータを書き込み
ws.append(["Month", "Total Sales"])  # ヘッダー
for row in monthly_sales.itertuples(index=False):
    ws.append(row)

# 5. 保存
output_file = "sales_data_updated.xlsx"
wb.save(output_file)

print(f"処理完了！結果は {output_file} に保存されました。")